from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter

from .parser import AnalysisResult


HEADER_FILL = PatternFill("solid", fgColor="07C160")
SUBTLE_FILL = PatternFill("solid", fgColor="F4FBF7")
WARN_FILL = PatternFill("solid", fgColor="FFE4E6")
TEXT = "1F2937"
GRID = Side(style="thin", color="D1D5DB")


def _style_sheet(ws) -> None:
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False
    border = Border(left=GRID, right=GRID, top=GRID, bottom=GRID)
    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center")
            if cell.row == 1:
                cell.fill = HEADER_FILL
                cell.font = Font(color="FFFFFF", bold=True)
            else:
                cell.font = Font(color=TEXT)
    for column in ws.columns:
        width = min(max(len(str(c.value or "")) for c in column) + 2, 36)
        ws.column_dimensions[get_column_letter(column[0].column)].width = max(width, 10)
    if ws.max_row > 1 and ws.max_column > 1:
        ws.auto_filter.ref = ws.dimensions


def _append_rows(ws, rows: list[list]) -> None:
    for row in rows:
        ws.append(row)
    _style_sheet(ws)


def export_excel(result: AnalysisResult, output_path: str | Path) -> Path:
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"

    summary_rows = [
        ["项目", "数值", "说明"],
        ["设备", result.device_name, f"SN {result.serial}"],
        ["数据范围", f"{result.range_start} 至 {result.range_end}", "DATALOG"],
        ["EDF 文件数", result.edf_count, ""],
        ["空 EDF", len(result.zero_edfs), "已跳过"],
        ["近一年有效天数", result.summary.get("year_days", 0), "STR.edf"],
        ["近一年平均使用", round(float(result.summary.get("year_avg_usage_hr", 0)), 2), "小时/晚"],
        ["近一年平均 AHI", round(float(result.summary.get("year_avg_ahi", 0)), 2), "次/小时"],
        ["近30天平均 AHI", round(float(result.summary.get("last30_avg_ahi", 0)), 2), "次/小时"],
        ["近30天 95%漏气", round(float(result.summary.get("last30_avg_leak95", 0)), 1), "L/min"],
        ["近30天漏气>=24", result.summary.get("last30_leak_over24_days", 0), "天"],
        ["当前压力范围", f"{result.summary.get('current_min_pressure', 0)} - {result.summary.get('current_max_pressure', 0)}", "cmH2O"],
        ["当前 EPR", result.summary.get("current_epr", 0), ""],
    ]
    _append_rows(ws, summary_rows)

    ws2 = wb.create_sheet("STR_Daily")
    _append_rows(
        ws2,
        [
            ["Date", "Usage_hr", "Duration_min", "AHI", "AI", "HI", "OAI", "CAI", "UAI", "Leak95_Lmin", "Leak50_Lmin", "LeakMax_Lmin", "Pressure95_cmH2O", "MinPressure", "MaxPressure", "EPR"],
            *[
                [d.date, round(d.usage_hr, 2), round(d.duration_min, 0), round(d.ahi, 2), round(d.ai, 2), round(d.hi, 2), round(d.oai, 2), round(d.cai, 2), round(d.uai, 2), round(d.leak95_lmin, 1), round(d.leak50_lmin, 1), round(d.leakmax_lmin, 1), round(d.pressure95, 2), d.min_pressure, d.max_pressure, d.epr_level]
                for d in result.str_days
            ],
        ],
    )
    ws2.conditional_formatting.add(f"J2:J{ws2.max_row}", CellIsRule(operator="greaterThanOrEqual", formula=["24"], fill=WARN_FILL))
    ws2.conditional_formatting.add(f"D2:D{ws2.max_row}", CellIsRule(operator="greaterThanOrEqual", formula=["5"], fill=PatternFill("solid", fgColor="FFF7ED")))

    ws3 = wb.create_sheet("DATALOG_Daily")
    _append_rows(
        ws3,
        [
            ["Date", "Usage_hr", "Duration_min", "Sessions", "TotalEvents", "Estimated_AHI", "CentralApnea", "ObstructiveApnea", "Hypopnea", "UnknownApnea", "RERA", "CSR_Events", "SkippedFiles"],
            *[
                [d.date, round(d.usage_hr, 2), round(d.duration_min, 0), d.sessions, d.total_events, round(d.estimated_ahi, 2), d.central_apnea, d.obstructive_apnea, d.hypopnea, d.apnea, d.rera, d.csr_events, d.skipped_files]
                for d in result.data_days
            ],
        ],
    )

    ws4 = wb.create_sheet("Leak_Watch")
    leak_days = [d for d in result.str_days if d.leak95_lmin >= 24]
    _append_rows(
        ws4,
        [
            ["Date", "Leak95_Lmin", "AHI", "Usage_hr", "Pressure95", "CAI", "OAI"],
            *[[d.date, round(d.leak95_lmin, 1), round(d.ahi, 2), round(d.usage_hr, 2), round(d.pressure95, 2), round(d.cai, 2), round(d.oai, 2)] for d in leak_days],
        ],
    )

    ws5 = wb.create_sheet("Suggestions")
    _append_rows(ws5, [["调整建议"], *[[x] for x in result.suggestions]])
    for row in range(2, ws5.max_row + 1):
        ws5.cell(row=row, column=1).fill = SUBTLE_FILL
        ws5.cell(row=row, column=1).alignment = Alignment(wrap_text=True, vertical="top")
    ws5.column_dimensions["A"].width = 100

    ws6 = wb.create_sheet("Codebook")
    _append_rows(
        ws6,
        [
            ["字段", "含义", "单位"],
            ["AHI", "呼吸暂停低通气指数", "次/小时"],
            ["CAI/OAI", "中枢/阻塞暂停指数", "次/小时"],
            ["Leak95_Lmin", "95%漏气", "L/min，24 为关注线"],
            ["Pressure95_cmH2O", "95%压力", "cmH2O"],
            ["Estimated_AHI", "DATALOG 事件重算 AHI", "估算值"],
        ],
    )

    wb.save(output)
    return output
